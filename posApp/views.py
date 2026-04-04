from datetime import datetime
from pickle import FALSE
import json
import sys
import io
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, JsonResponse
from posApp.models import Category, Employee, Products, Sales, salesItems, Customer, WeeklyDebit, Attendance
from django.db.models import Count, Sum
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt
import json, sys
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth.models import User, Group
from django.views.decorators.cache import cache_control
from django.utils.timezone import localtime, now
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import timedelta, date
from django.core.mail import EmailMessage


# Login
def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Incorrect username or password"
        else:
            resp['msg'] = "Incorrect username or password"
    return HttpResponse(json.dumps(resp),content_type='application/json')

#Logout
def logoutuser(request):
    logout(request)
    return redirect('/')

# Create your views here.
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def home(request):
    user_groups = request.user.groups.all()
    u = request.user

    # Get the current date and time in the local timezone
    current_date = localtime(now()).date()

    # Total counts for categories, products, and transactions
    categories = Category.objects.count()
    products = Products.objects.count()

    # Filter transactions based on the current date
    today_sales = Sales.objects.filter(date_added__date=current_date)

    transaction = today_sales.count()

    # Sales for today
    total_sales = today_sales.aggregate(total=Sum('grand_total'))['total'] or 0
    total_sales_cash = today_sales.filter(payment_method='cash').aggregate(total=Sum('grand_total'))['total'] or 0
    total_sales_card = today_sales.filter(payment_method='card').aggregate(total=Sum('grand_total'))['total'] or 0
    total_sales_gpay = today_sales.filter(payment_method='gpay').aggregate(total=Sum('grand_total'))['total'] or 0

    room_1_attendance = today_sales.filter(room_no=1).count()
    room_2_attendance = today_sales.filter(room_no=2).count()
    room_3_attendance = today_sales.filter(room_no=3).count()
    room_4_attendance = today_sales.filter(room_no=4).count()

    # Token counts by room and product
    room_1_token_counts = (
        today_sales.filter(room_no=1)
        .values('items__product_id__name')
        .annotate(token_count=Count('token_no'))
        .order_by('items__product_id__name')
    )

    room_2_token_counts = (
        today_sales.filter(room_no=2)
        .values('items__product_id__name')
        .annotate(token_count=Count('token_no'))
        .order_by('items__product_id__name')
    )

    room_3_token_counts = (
        today_sales.filter(room_no=3)
        .values('items__product_id__name')
        .annotate(token_count=Count('token_no'))
        .order_by('items__product_id__name')
    )

    room_4_token_counts = (
        today_sales.filter(room_no=4)
        .values('items__product_id__name')
        .annotate(token_count=Count('token_no'))
        .order_by('items__product_id__name')
    )


    # Latest sale token number
    token_counts_by_product = (
        today_sales
        .values('items__product_id__name')  # Group by product name
        .annotate(token_count=Count('token_no'))  # Count tokens per product
        .order_by('items__product_id__name')  # Sort by product name
    )

    # Latest sale token number
    latest_sale = today_sales.order_by('-token_no').first()
    token_no = latest_sale.token_no if latest_sale else 0


    # Context to pass to the template
    context = {
        'user_groups': user_groups,
        'u': u,
        'page_title': 'Home',
        'categories': categories,
        'products': products,
        'transaction': transaction,
        'total_sales': total_sales,
        'total_sales_cash': total_sales_cash,
        'total_sales_card': total_sales_card,
        'total_sales_gpay': total_sales_gpay,
        'token_no': token_no,
        'token_counts_by_product': token_counts_by_product,
        'room_1_attendance': room_1_attendance,
        'room_2_attendance': room_2_attendance,
        'room_3_attendance': room_3_attendance,
        'room_4_attendance': room_4_attendance,
        'room_1_token_counts': room_1_token_counts,
        'room_2_token_counts': room_2_token_counts,
        'room_3_token_counts': room_3_token_counts,
        'room_4_token_counts': room_4_token_counts,
    }
    
    return render(request, 'posApp/home.html', context)


def about(request):
    context = {
        'page_title':'About',
    }
    return render(request, 'posApp/about.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def add_user(request):
    user_groups = request.user.groups.all()
    u = request.user

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        password_confirm = request.POST['password_confirm']
        email = request.POST['email']
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        group_id = request.POST.get('group', '')

        if password != password_confirm:
            messages.error(request, 'Passwords do not match.')
            return redirect('add_user')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('add_user')

        user = User.objects.create_user(username=username, password=password, email=email,
                                        first_name=first_name, last_name=last_name)

        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                user.groups.add(group)
            except Group.DoesNotExist:
                messages.error(request, 'Selected group does not exist.')
                return redirect('add_user')

        messages.success(request, 'User created successfully.')
        return redirect('add_user')

    groups = Group.objects.all()
    context = {
        'groups': groups,
        'user_groups': user_groups,
        'u': u,
    }
    
    return render(request, 'posApp/add_user.html', context)

#Categories
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def category(request):
    user_groups = request.user.groups.all()
    u = request.user

    category_list = Category.objects.all()
    # category_list = {}
    context = {
        'page_title':'Category List',
        'category':category_list,
        'user_groups': user_groups,
        'u': u,
    }
    return render(request, 'posApp/category.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def manage_category(request):
    user_groups = request.user.groups.all()
    u = request.user

    category = {}
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            category = Category.objects.filter(id=id).first()
    
    context = {
        'category' : category,
        'user_groups': user_groups,
        'u': u,
    }
    return render(request, 'posApp/manage_category.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def save_category(request):
    data =  request.POST
    resp = {'status':'failed'}
    try:
        if (data['id']).isnumeric() and int(data['id']) > 0 :
            save_category = Category.objects.filter(id = data['id']).update(name=data['name'],status = data['status'])
        else:
            save_category = Category(name=data['name'],status = data['status'])
            save_category.save()
        resp['status'] = 'success'
        messages.success(request, 'Category Successfully saved.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def delete_category(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Category.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
        messages.success(request, 'Category Successfully deleted.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

# Products
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def products(request):
    user_groups = request.user.groups.all()
    u = request.user

    product_list = Products.objects.all()
    context = {
        'page_title':'Product List',
        'products':product_list,
        'user_groups': user_groups,
        'u': u,
    }
    return render(request, 'posApp/products.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def manage_products(request):
    user_groups = request.user.groups.all()
    u = request.user

    product = {}
    categories = Category.objects.filter(status = 1).all()
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            product = Products.objects.filter(id=id).first()
    
    context = {
        'product' : product,
        'categories' : categories,
        'user_groups': user_groups,
        'u': u,
    }
    return render(request, 'posApp/manage_product.html',context)

def test(request):
    categories = Category.objects.all()
    context = {
        'categories' : categories
    }
    return render(request, 'posApp/test.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def save_product(request):
    data = request.POST
    resp = {'status': 'failed'}
    id = ''
    
    if 'id' in data:
        id = data['id']
        
    if id.isnumeric() and int(id) > 0:
        check = Products.objects.exclude(id=id).filter(name=data['name']).all()  # Update check based on name or other field if needed
    else:
        check = Products.objects.filter(name=data['name']).all()  # Update check based on name or other field if needed
    
    if len(check) > 0:
        resp['msg'] = "Product with this name already exists in the database"
    else:
        category = Category.objects.filter(id=data['category_id']).first()
        try:
            if id.isnumeric() and int(id) > 0:
                Products.objects.filter(id=id).update(
                    category_id=category,
                    name=data['name'],
                    price=float(data['price']),
                    status=data['status']
                )
            else:
                new_product = Products(
                    category_id=category,
                    name=data['name'],
                    price=float(data['price']),
                    status=data['status']
                )
                new_product.save()
            resp['status'] = 'success'
            messages.success(request, 'Product successfully saved.')
        except Exception as e:
            resp['status'] = 'failed'
            resp['error'] = str(e)
    
    return HttpResponse(json.dumps(resp), content_type="application/json")

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def delete_product(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Products.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
        messages.success(request, 'Product Successfully deleted.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def pos(request):
    user_groups = request.user.groups.all()
    u = request.user

    products = Products.objects.filter(status = 1)
    product_json = []
    for product in products:
        product_json.append({'id':product.id, 'name':product.name, 'price':float(product.price)})
    context = {
        'page_title' : "Point of Sale",
        'products' : products,
        'product_json' : json.dumps(product_json),
        'user_groups': user_groups,
        'u': u,
    }
    # return HttpResponse('')
    return render(request, 'posApp/pos.html',context)

# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# @login_required()
# def checkout_modal(request):
#     user_groups = request.user.groups.all()
#     u = request.user

#     grand_total = 0
#     if 'grand_total' in request.GET:
#         grand_total = request.GET['grand_total']
#     context = {
#         'grand_total' : grand_total,
#         'user_groups': user_groups,
#         'u': u,
#     }
#     return render(request, 'posApp/checkout.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def save_pos(request):
    resp = {'status': 'failed', 'msg': ''}
    data = request.POST

    try:
        # Check if customer exists in the database by phone number
        customer, created = Customer.objects.get_or_create(
            phone_number=data['customer_phone'],
            defaults={
                'name': data['customer_name'],
                'city': data['customer_city']
            }
        )

        if not created:
            # If customer already exists, update the name and city
            customer.name = data['customer_name']
            customer.city = data['customer_city']
            customer.save()

        # Save Sale Record
        sales = Sales(
            customer_name=data['customer_name'],
            customer_phone=data['customer_phone'],
            customer_city=data['customer_city'],
            room_no=data['room_no'],
            payment_method=data['payment_method'],
            sub_total=data['sub_total'],
            grand_total=data['grand_total'],
            date_added=timezone.now()  # Ensure the correct current date is used
        )
        sales.save()  # This will trigger the `save` method in the Sales model

        sale_id = sales.pk
        for i, prod in enumerate(data.getlist('product_id[]')):
            product = Products.objects.get(id=prod)
            qty = data.getlist('qty[]')[i]
            price = data.getlist('price[]')[i]
            total = float(qty) * float(price)

            salesItems(sale_id=sales, product_id=product, qty=qty, price=price, total=total).save()

        resp['status'] = 'success'
        resp['sale_id'] = sale_id
        
        # messages.success(request, "Sale Record has been saved.")
    except Exception as e:
        resp['msg'] = "An error occurred"
        print(e)  # For debugging purposes

    return JsonResponse(resp)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def get_customer_by_phone(request):
    phone_number = request.GET.get('customer_phone', None)
    if phone_number:
        try:
            customer = Customer.objects.get(phone_number=phone_number)
            # If customer exists, return their details
            return JsonResponse({
                'status': 'success',
                'customer_name': customer.name,
                'customer_city': customer.city
            })
        except Customer.DoesNotExist:
            return JsonResponse({
                'status': 'not_found',
                'msg': 'Customer not found'
            })
    else:
        return JsonResponse({'status': 'failed', 'msg': 'No phone number provided'})

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def salesList(request):
    user_groups = request.user.groups.all()
    u = request.user

    search_query = request.GET.get('q', '')
    date_str = request.GET.get('date', '')

    date = datetime.strptime(date_str, '%Y-%m-%d') if date_str else None

    # Initialize `sales` with all Sales objects
    sales = Sales.objects.all()

    if search_query:
        sales = sales.filter(
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(token_no__icontains=search_query) |
            Q(room_no__icontains=search_query)   
        )
    
    if date:
        sales = sales.filter(date_added__date=date)
    
    # Apply ordering to the filtered queryset
    sales = sales.order_by('-date_added')
    
    sale_data = []
    for sale in sales:
        data = {}
        for field in sale._meta.get_fields(include_parents=False):
            if field.related_model is None:
                data[field.name] = getattr(sale, field.name)
        data['items'] = salesItems.objects.filter(sale_id=sale).all()
        data['item_count'] = len(data['items'])
        if 'tax_amount' in data:
            data['tax_amount'] = format(float(data['tax_amount']), '.2f')
        sale_data.append(data)

    # Calculate products sold count
    products_sold_count = salesItems.objects.filter(sale_id__in=sales).values('product_id__name').annotate(total_sold=Sum('qty')).order_by('-total_sold')

    # Check if download is requested
    if 'download' in request.GET:
        return export_sales_to_excel(sale_data)

    context = {
        'page_title': 'Sales Transactions',
        'sale_data': sale_data,
        'user_groups': user_groups,
        'u': u,
        'search_query': search_query,
        'date': date_str,
        'products_sold_count': products_sold_count,
    }

    return render(request, 'posApp/sales.html', context)

def export_sales_to_excel(sale_data):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # Set headers for the sheet
    headers = ['S.No', 'DateTime', 'Customer Name', 'Token No', 'Total', 'Items']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Add data to the sheet
    for row_num, sale in enumerate(sale_data, 2):
        ws[f'A{row_num}'] = row_num - 1  # S.No
        ws[f'B{row_num}'] = sale['date_added'].strftime('%d-%m-%Y, %I:%M %p')  # DateTime
        ws[f'C{row_num}'] = sale['customer_name']  # Customer Name
        ws[f'D{row_num}'] = sale['token_no']  # Token No
        ws[f'E{row_num}'] = sale['grand_total']  # Total
        ws[f'F{row_num}'] = sale['item_count']  # Items

    # Set up the response for downloading the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_data.xlsx"'
    wb.save(response)
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def receipt(request):
    user_groups = request.user.groups.all()
    u = request.user

    id = request.GET.get('id')
    sales = Sales.objects.filter(id = id).first()
    transaction = {}
    for field in Sales._meta.get_fields():
        if field.related_model is None:
            transaction[field.name] = getattr(sales,field.name)
    if 'tax_amount' in transaction:
        transaction['tax_amount'] = format(float(transaction['tax_amount']))
    ItemList = salesItems.objects.filter(sale_id = sales).all()
    context = {
        "transaction" : transaction,
        "salesItems" : ItemList,
        'user_groups': user_groups,
        'u': u,
    }

    return render(request, 'posApp/receipt.html',context)
    # return HttpResponse('')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def delete_sale(request):
    resp = {'status':'failed', 'msg':''}
    id = request.POST.get('id')
    try:
        delete = Sales.objects.filter(id = id).delete()
        resp['status'] = 'success'
        messages.success(request, 'Sale Record has been deleted.')
    except:
        resp['msg'] = "An error occured"
        print("Unexpected error:", sys.exc_info()[0])
    return HttpResponse(json.dumps(resp), content_type='application/json')

def delete_multiple_sales(request):
    resp = {'status': 'failed', 'msg': ''}
    try:
        # Parse JSON payload from the request body
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        print("IDs received for deletion:", ids)  # Debug: Check received IDs
        
        # Proceed with deletion if IDs are provided
        if ids:
            delete = Sales.objects.filter(id__in=ids).delete()
            if delete[0] > 0:
                resp['status'] = 'success'
                messages.success(request, 'Selected Sale Records have been deleted.')
            else:
                resp['msg'] = "No records were deleted."
        else:
            resp['msg'] = "No IDs provided for deletion."
    except Exception as e:
        resp['msg'] = "An error occurred"
        print("Unexpected error:", e)
    
    return JsonResponse(resp)



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def customerdatabase(request):
    user_groups = request.user.groups.all()
    u = request.user
    query = request.GET.get('q')
    customers = Customer.objects.all()

    if query:
        customers = customers.filter(
            Q(name__icontains=query) | 
            Q(phone_number__icontains=query) |  
            Q(city__icontains=query)
        )


    context = {
        'user_groups': user_groups,
        'u': u,
        'customers': customers,
        'search_query': query,
    }

    return render(request, 'posApp/customerdatabase.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def delete_customer(request):
    customer_id = request.GET.get('id')
    if customer_id:
        try:
            customer = Customer.objects.get(pk=customer_id)
            customer.delete()
            return JsonResponse({'success': True})
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Customer does not exist.'})
    return JsonResponse({'success': False, 'error': 'Invalid customer ID.'})

def get_sunday(d):
    return d - timedelta(days=d.weekday() + 1) if d.weekday() != 6 else d

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def report(request):
    # Get filter date from request or default to today
    filter_date_str = request.GET.get('filter_date')
    if filter_date_str:
        try:
            # Parse the date from the input
            year, month, day = map(int, filter_date_str.split('-'))
            filter_date = date(year, month, day)
        except (ValueError, TypeError):
            filter_date = timezone.now().date()
    else:
        filter_date = timezone.now().date()
    
    today = filter_date
    first_sunday = get_sunday(date(today.year, 1, 1))
    current_sunday = get_sunday(today)

    weekly_reports = []
    week_number = 1
    while first_sunday <= current_sunday:
        week_start = first_sunday
        week_end = week_start + timedelta(days=6)
        is_current_week = week_start <= today <= week_end

        # Check if the selected date falls within this week
        is_selected_week = week_start <= filter_date <= week_end

        week_sales = Sales.objects.filter(date_added__date__gte=week_start, date_added__date__lte=week_end)
        total_sales = week_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        total_sales = float(total_sales)  # Convert to float

        # Get debits for the week
        debits = WeeklyDebit.objects.filter(week_number=week_number)
        total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
        total_debits = float(total_debits)  # Convert to float
        
        # Calculate net sales
        net_sales = total_sales - total_debits

        # Include in reports if no filter is applied or if the week matches the filter
        if not filter_date_str or is_selected_week:
            weekly_reports.append({
                'week_number': week_number,
                'start_date': week_start.strftime("%d-%b-%Y"),
                'end_date': week_end.strftime("%d-%b-%Y"),
                'total_sales': total_sales,
                'total_debits': total_debits,
                'net_sales': net_sales,
                'is_current_week': is_current_week,
                'is_selected_week': is_selected_week
            })

        first_sunday += timedelta(days=7)
        week_number += 1

    weekly_reports.reverse()

    context = {
        'weekly_reports': weekly_reports,
        'user_groups': request.user.groups.all(),
        'u': request.user,
        'filter_date': filter_date,
        'today': timezone.now().date(),
        'total_sales': '{:,.2f}'.format(total_sales) if 'total_sales' in locals() else '{:,.2f}'.format(0),
    }
    return render(request, 'posApp/report.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def weekly_report_detail(request, week_number):
    today = timezone.now().date()
    first_sunday = get_sunday(date(today.year, 1, 1))
    
    # Calculate the start date for the requested week
    week_start = first_sunday + timedelta(days=7 * (week_number - 1))
    week_end = week_start + timedelta(days=6)
    
    # Get sales for the week
    week_sales = Sales.objects.filter(
        date_added__date__gte=week_start, 
        date_added__date__lte=week_end
    )
    
    # Calculate total sales for the week
    total_sales = week_sales.aggregate(total=Sum('grand_total'))['total'] or 0
    total_sales = float(total_sales)  # Convert to float
    
    # Get debits for the week
    from posApp.models import WeeklyDebit
    debits = WeeklyDebit.objects.filter(week_number=week_number)
    total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
    total_debits = float(total_debits)  # Convert to float
    
    # Calculate net sales
    net_sales = total_sales - total_debits
    
    # Calculate daily sales - starting with Sunday
    day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    daily_sales = []
    
    for i in range(7):
        current_date = week_start + timedelta(days=i)
        day_sales = Sales.objects.filter(date_added__date=current_date)
        day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        day_total = float(day_total)  # Convert to float
        
        daily_sales.append({
            'day': day_names[i],  # Use predefined day names
            'date': current_date.strftime('%d-%b-%Y'),
            'sales': day_total,
        })
    
    context = {
        'week_number': week_number,
        'week_start': week_start.strftime('%A, %B %d, %Y'),
        'week_end': week_end.strftime('%A, %B %d, %Y'),
        'total_sales': total_sales,
        'total_debits': total_debits,
        'net_sales': net_sales,
        'daily_sales': daily_sales,
        'debits': debits,
        'user_groups': request.user.groups.all(),
        'u': request.user,
        'page_title': f'Report for Week {week_number}'
    }
    
    return render(request, 'posApp/weekly_report_detail.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def add_weekly_debit(request, week_number):
    if request.method == 'POST':
        description = request.POST.get('description')
        amount = request.POST.get('amount')
        
        if description and amount:
            from posApp.models import WeeklyDebit
            
            # Calculate the start date for the requested week
            today = timezone.now().date()
            first_sunday = get_sunday(date(today.year, 1, 1))
            week_start = first_sunday + timedelta(days=7 * (week_number - 1))
            
            # Create new debit - convert amount to float
            try:
                amount = float(amount)
                # Create new debit entry
                WeeklyDebit.objects.create(
                    week_number=week_number,
                    week_start_date=week_start,
                    description=description,
                    amount=amount
                )
                messages.success(request, 'Debit added successfully')
            except ValueError:
                messages.error(request, 'Invalid amount entered')
                return redirect('weekly_report_detail', week_number=week_number)
            messages.success(request, 'Debit added successfully!')
        else:
            messages.error(request, 'Description and amount are required!')
            
    return redirect('weekly_report_detail', week_number=week_number)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def edit_weekly_debit(request, pk):
    debit = get_object_or_404(WeeklyDebit, pk=pk)
    week_number = debit.week_number
    
    if request.method == 'POST':
        description = request.POST.get('description')
        amount = request.POST.get('amount')
        
        if description and amount:
            try:
                amount = float(amount)
                # Update debit entry
                debit.description = description
                debit.amount = amount
                debit.save()
                messages.success(request, 'Debit updated successfully')
            except ValueError:
                messages.error(request, 'Invalid amount entered')
        else:
            messages.error(request, 'Description and amount are required!')
            
    return redirect('weekly_report_detail', week_number=week_number)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def delete_weekly_debit(request, pk):
    debit = get_object_or_404(WeeklyDebit, pk=pk)
    week_number = debit.week_number
    
    debit.delete()
    messages.success(request, 'Debit deleted successfully')
    
    return redirect('weekly_report_detail', week_number=week_number)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def export_weekly_report(request, week_number):
    if request.method == 'POST':
        include_daily_sales = 'include_daily_sales' in request.POST
        include_debits = 'include_debits' in request.POST
        
        # Get data for the week
        today = timezone.now().date()
        first_sunday = get_sunday(date(today.year, 1, 1))
        week_start = first_sunday + timedelta(days=7 * (week_number - 1))
        week_end = week_start + timedelta(days=6)
        
        # Get sales data
        week_sales = Sales.objects.filter(date_added__date__gte=week_start, date_added__date__lte=week_end)
        total_sales = week_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        total_sales = float(total_sales)
        
        # Get debit data
        debits = WeeklyDebit.objects.filter(week_number=week_number)
        total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
        total_debits = float(total_debits)
        
        # Calculate net sales
        net_sales = total_sales - total_debits
        
        # Calculate daily sales
        day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        daily_sales = []
        
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            day_sales = Sales.objects.filter(date_added__date=current_date)
            day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            day_total = float(day_total)
            
            daily_sales.append({
                'day': day_names[i],
                'date': current_date.strftime('%d-%b-%Y'),
                'sales': day_total,
            })
        
        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        
        # Add title
        ws['A1'] = f"Weekly Sales Report - Week {week_number}"
        ws['A2'] = f"Period: {week_start.strftime('%d-%b-%Y')} to {week_end.strftime('%d-%b-%Y')}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Add summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = openpyxl.styles.Font(bold=True)
        
        ws['A5'] = "Gross Sales:"
        ws['B5'] = total_sales
        ws['B5'].number_format = '₹#,##0.00'
        
        ws['A6'] = "Total Debits:"
        ws['B6'] = total_debits
        ws['B6'].number_format = '₹#,##0.00'
        
        ws['A7'] = "Net Balance:"
        ws['B7'] = net_sales
        ws['B7'].font = openpyxl.styles.Font(bold=True)
        ws['B7'].number_format = '₹#,##0.00'
        
        row = 9  # Start row for the next section
        
        # Add daily sales if selected
        if include_daily_sales:
            ws[f'A{row}'] = "DAILY SALES"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add headers
            ws[f'A{row}'] = "Day"
            ws[f'B{row}'] = "Date"
            ws[f'C{row}'] = "Sales Amount"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add data
            for day_data in daily_sales:
                ws[f'A{row}'] = day_data['day']
                ws[f'B{row}'] = day_data['date']
                ws[f'C{row}'] = day_data['sales']
                ws[f'C{row}'].number_format = '₹#,##0.00'
                row += 1
            
            row += 2  # Add space after daily sales
        
        # Add debit history if selected
        if include_debits:
            ws[f'A{row}'] = "DEBIT HISTORY"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            if debits.exists():
                # Add headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Description"
                ws[f'C{row}'] = "Amount"
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add data
                for debit in debits:
                    ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                    ws[f'B{row}'] = debit.description
                    ws[f'C{row}'] = debit.amount
                    ws[f'C{row}'].number_format = '₹#,##0.00'
                    row += 1
            else:
                ws[f'A{row}'] = "No debits recorded for this week."
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set up the response for downloading the file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Weekly_Report_Week_{week_number}_{week_start.strftime('%d_%b_%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    return redirect('weekly_report_detail', week_number=week_number)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def export_monthly_report(request):
    if request.method == 'POST':
        month_str = request.POST.get('month')  # Format: YYYY-MM
        include_daily_sales = 'include_daily_sales' in request.POST
        include_debits = 'include_debits' in request.POST
        
        # Parse month and year
        try:
            year, month = map(int, month_str.split('-'))
            # Get first and last day of the month
            first_day = date(year, month, 1)
            # Get last day by finding first day of next month and subtracting one day
            if month == 12:
                last_day = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = date(year, month + 1, 1) - timedelta(days=1)
        except (ValueError, AttributeError, TypeError):
            messages.error(request, 'Invalid month selected')
            return redirect('report')
        
        # Get sales data for the month
        month_sales = Sales.objects.filter(date_added__date__gte=first_day, date_added__date__lte=last_day)
        total_sales = month_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        total_sales = float(total_sales)
        
        # Get debit data for the month
        # First, get all week numbers that fall within this month
        relevant_weeks = set()
        current_date = first_day
        while current_date <= last_day:
            first_sunday_of_year = get_sunday(date(year, 1, 1))
            days_since_first_sunday = (current_date - first_sunday_of_year).days
            week_number = (days_since_first_sunday // 7) + 1
            relevant_weeks.add(week_number)
            current_date += timedelta(days=1)
        
        # Get all debits for these weeks
        debits = WeeklyDebit.objects.filter(week_number__in=relevant_weeks)
        total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
        total_debits = float(total_debits)
        
        # Calculate net sales
        net_sales = total_sales - total_debits
        
        # Calculate daily sales
        daily_sales = []
        current_date = first_day
        while current_date <= last_day:
            day_sales = Sales.objects.filter(date_added__date=current_date)
            day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            day_total = float(day_total)
            
            daily_sales.append({
                'date': current_date.strftime('%d-%b-%Y'),
                'day': current_date.strftime('%a'),
                'sales': day_total,
            })
            current_date += timedelta(days=1)
        
        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        
        # Add title
        month_name = first_day.strftime('%B')
        ws['A1'] = f"Monthly Sales Report - {month_name} {year}"
        ws['A2'] = f"Period: {first_day.strftime('%d-%b-%Y')} to {last_day.strftime('%d-%b-%Y')}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Add summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = openpyxl.styles.Font(bold=True)
        
        ws['A5'] = "Gross Sales:"
        ws['B5'] = total_sales
        ws['B5'].number_format = '₹#,##0.00'
        
        ws['A6'] = "Total Debits:"
        ws['B6'] = total_debits
        ws['B6'].number_format = '₹#,##0.00'
        
        ws['A7'] = "Net Balance:"
        ws['B7'] = net_sales
        ws['B7'].font = openpyxl.styles.Font(bold=True)
        ws['B7'].number_format = '₹#,##0.00'
        
        row = 9  # Start row for the next section
        
        # Add daily sales if selected
        if include_daily_sales:
            ws[f'A{row}'] = "DAILY SALES"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add headers
            ws[f'A{row}'] = "Date"
            ws[f'B{row}'] = "Day"
            ws[f'C{row}'] = "Sales Amount"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add data
            for day_data in daily_sales:
                ws[f'A{row}'] = day_data['date']
                ws[f'B{row}'] = day_data['day']
                ws[f'C{row}'] = day_data['sales']
                ws[f'C{row}'].number_format = '₹#,##0.00'
                row += 1
            
            row += 2  # Add space after daily sales
        
        # Add debit history if selected
        if include_debits:
            ws[f'A{row}'] = "DEBIT HISTORY"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            if debits.exists():
                # Add headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Week #"
                ws[f'C{row}'] = "Description"
                ws[f'D{row}'] = "Amount"
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add data
                for debit in debits:
                    ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                    ws[f'B{row}'] = f"Week {debit.week_number}"
                    ws[f'C{row}'] = debit.description
                    ws[f'D{row}'] = debit.amount
                    ws[f'D{row}'].number_format = '₹#,##0.00'
                    row += 1
            else:
                ws[f'A{row}'] = "No debits recorded for this month."
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set up the response for downloading the file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Monthly_Report_{month_name}_{year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    messages.error(request, 'Invalid request method')
    return redirect('report')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def email_monthly_report(request):
    if request.method == 'POST':
        month_str = request.POST.get('month')  # Format: YYYY-MM
        recipient_email = request.POST.get('recipient_email', 'sivaprasanth8208@gmail.com')
        include_daily_sales = 'include_daily_sales' in request.POST
        include_debits = 'include_debits' in request.POST
        
        # Parse month and year
        try:
            year, month = map(int, month_str.split('-'))
            # Get first and last day of the month
            first_day = date(year, month, 1)
            # Get last day by finding first day of next month and subtracting one day
            if month == 12:
                last_day = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = date(year, month + 1, 1) - timedelta(days=1)
        except (ValueError, AttributeError, TypeError):
            messages.error(request, 'Invalid month selected')
            return redirect('report')
        
        # Get sales data for the month
        month_sales = Sales.objects.filter(date_added__date__gte=first_day, date_added__date__lte=last_day)
        total_sales = month_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        total_sales = float(total_sales)
        
        # Get debit data for the month
        # First, get all week numbers that fall within this month
        relevant_weeks = set()
        current_date = first_day
        while current_date <= last_day:
            first_sunday_of_year = get_sunday(date(year, 1, 1))
            days_since_first_sunday = (current_date - first_sunday_of_year).days
            week_number = (days_since_first_sunday // 7) + 1
            relevant_weeks.add(week_number)
            current_date += timedelta(days=1)
        
        # Get all debits for these weeks
        debits = WeeklyDebit.objects.filter(week_number__in=relevant_weeks)
        total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
        total_debits = float(total_debits)
        
        # Calculate net sales
        net_sales = total_sales - total_debits
        
        # Calculate daily sales
        daily_sales = []
        current_date = first_day
        while current_date <= last_day:
            day_sales = Sales.objects.filter(date_added__date=current_date)
            day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            day_total = float(day_total)
            
            daily_sales.append({
                'date': current_date.strftime('%d-%b-%Y'),
                'day': current_date.strftime('%a'),
                'sales': day_total,
            })
            current_date += timedelta(days=1)
        
        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        
        # Add title
        month_name = first_day.strftime('%B')
        ws['A1'] = f"Monthly Sales Report - {month_name} {year}"
        ws['A2'] = f"Period: {first_day.strftime('%d-%b-%Y')} to {last_day.strftime('%d-%b-%Y')}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Add summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = openpyxl.styles.Font(bold=True)
        
        ws['A5'] = "Gross Sales:"
        ws['B5'] = total_sales
        ws['B5'].number_format = '₹#,##0.00'
        
        ws['A6'] = "Total Debits:"
        ws['B6'] = total_debits
        ws['B6'].number_format = '₹#,##0.00'
        
        ws['A7'] = "Net Balance:"
        ws['B7'] = net_sales
        ws['B7'].font = openpyxl.styles.Font(bold=True)
        ws['B7'].number_format = '₹#,##0.00'
        
        row = 9  # Start row for the next section
        
        # Add daily sales if selected
        if include_daily_sales:
            ws[f'A{row}'] = "DAILY SALES"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add headers
            ws[f'A{row}'] = "Date"
            ws[f'B{row}'] = "Day"
            ws[f'C{row}'] = "Sales Amount"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add data
            for day_data in daily_sales:
                ws[f'A{row}'] = day_data['date']
                ws[f'B{row}'] = day_data['day']
                ws[f'C{row}'] = day_data['sales']
                ws[f'C{row}'].number_format = '₹#,##0.00'
                row += 1
            
            row += 2  # Add space after daily sales
        
        # Add debit history if selected
        if include_debits:
            ws[f'A{row}'] = "DEBIT HISTORY"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            if debits.exists():
                # Add headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Week #"
                ws[f'C{row}'] = "Description"
                ws[f'D{row}'] = "Amount"
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add data
                for debit in debits:
                    ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                    ws[f'B{row}'] = f"Week {debit.week_number}"
                    ws[f'C{row}'] = debit.description
                    ws[f'D{row}'] = debit.amount
                    ws[f'D{row}'].number_format = '₹#,##0.00'
                    row += 1
            else:
                ws[f'A{row}'] = "No debits recorded for this month."
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to a temporary file
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create the email
        email = EmailMessage(
            subject=f"Monthly Report - {month_name} {year}",
            body=f"Please find attached the monthly sales report for {month_name} {year}.\n\n"
                 f"Period: {first_day.strftime('%d-%b-%Y')} to {last_day.strftime('%d-%b-%Y')}\n"
                 f"Gross Sales: ₹{total_sales:,.2f}\n"
                 f"Total Debits: ₹{total_debits:,.2f}\n"
                 f"Net Balance: ₹{net_sales:,.2f}",
            from_email="system@bavaastro.com",  # Update with your email
            to=[recipient_email]
        )
        
        # Attach the Excel file
        email.attach(f"Monthly_Report_{month_name}_{year}.xlsx", excel_file.getvalue(), 
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Send the email
        try:
            email.send()
            messages.success(request, f'Monthly report for {month_name} {year} has been emailed to {recipient_email}')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')
        
        return redirect('report')
    
    messages.error(request, 'Invalid request method')
    return redirect('report')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def export_date_range_report(request):
    if request.method == 'POST':
        try:
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            include_daily_sales = 'include_daily_sales' in request.POST
            include_debits = 'include_debits' in request.POST
            
            # Parse the dates
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Validate date range
            if start_date > end_date:
                messages.error(request, 'Start date cannot be after end date')
                return redirect('report')
                
            # Get sales data for the date range
            range_sales = Sales.objects.filter(date_added__date__gte=start_date, date_added__date__lte=end_date)
            total_sales = range_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            total_sales = float(total_sales)
            
            # Get all week numbers that fall within this date range
            relevant_weeks = set()
            current_date = start_date
            while current_date <= end_date:
                year = current_date.year
                first_sunday_of_year = get_sunday(date(year, 1, 1))
                days_since_first_sunday = (current_date - first_sunday_of_year).days
                week_number = (days_since_first_sunday // 7) + 1
                relevant_weeks.add(week_number)
                current_date += timedelta(days=1)
            
            # Get all debits for these weeks
            debits = WeeklyDebit.objects.filter(week_number__in=relevant_weeks)
            total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
            total_debits = float(total_debits)
            
            # Calculate net sales
            net_sales = total_sales - total_debits
            
            # Calculate daily sales
            daily_sales = []
            current_date = start_date
            while current_date <= end_date:
                day_sales = Sales.objects.filter(date_added__date=current_date)
                day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
                day_total = float(day_total)
                
                daily_sales.append({
                    'date': current_date.strftime('%d-%b-%Y'),
                    'day': current_date.strftime('%a'),
                    'sales': day_total,
                })
                current_date += timedelta(days=1)
            
            # Create an Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Date Range Report"
            
            # Add title
            ws['A1'] = f"Sales Report - {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
            ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            
            # Add date range
            date_range_days = (end_date - start_date).days + 1
            ws['A2'] = f"Period: {date_range_days} days"
            ws['A2'].font = openpyxl.styles.Font(italic=True)
            
            # Add summary
            ws['A4'] = "SUMMARY"
            ws['A4'].font = openpyxl.styles.Font(bold=True)
            
            ws['A5'] = "Gross Sales:"
            ws['B5'] = total_sales
            ws['B5'].number_format = '₹#,##0.00'
            
            ws['A6'] = "Total Debits:"
            ws['B6'] = total_debits
            ws['B6'].number_format = '₹#,##0.00'
            
            ws['A7'] = "Net Balance:"
            ws['B7'] = net_sales
            ws['B7'].font = openpyxl.styles.Font(bold=True)
            ws['B7'].number_format = '₹#,##0.00'
            
            row = 9  # Start row for the next section
            
            # Add daily sales if selected
            if include_daily_sales:
                ws[f'A{row}'] = "DAILY SALES"
                ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Day"
                ws[f'C{row}'] = "Sales Amount"
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add data
                for day_data in daily_sales:
                    ws[f'A{row}'] = day_data['date']
                    ws[f'B{row}'] = day_data['day']
                    ws[f'C{row}'] = day_data['sales']
                    ws[f'C{row}'].number_format = '₹#,##0.00'
                    row += 1
                
                row += 2  # Add space after daily sales
            
            # Add debit history if selected
            if include_debits:
                ws[f'A{row}'] = "DEBIT HISTORY"
                ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                if debits.exists():
                    # Add headers
                    ws[f'A{row}'] = "Date"
                    ws[f'B{row}'] = "Week #"
                    ws[f'C{row}'] = "Description"
                    ws[f'D{row}'] = "Amount"
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.font = openpyxl.styles.Font(bold=True)
                    row += 1
                    
                    # Add data
                    for debit in debits:
                        ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                        ws[f'B{row}'] = f"Week {debit.week_number}"
                        ws[f'C{row}'] = debit.description
                        ws[f'D{row}'] = debit.amount
                        ws[f'D{row}'].number_format = '₹#,##0.00'
                        row += 1
                else:
                    ws[f'A{row}'] = "No debits recorded for this date range."
                    row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set up the response for downloading the file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Sales_Report_{start_date.strftime('%d_%b_%Y')}_to_{end_date.strftime('%d_%b_%Y')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            
            return response
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid date format: {str(e)}')
            return redirect('report')
    
    messages.error(request, 'Invalid request method')
    return redirect('report')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def email_date_range_report(request):
    if request.method == 'POST':
        try:
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            recipient_email = request.POST.get('recipient_email', 'sivaprasanth8208@gmail.com')
            include_daily_sales = 'include_daily_sales' in request.POST
            include_debits = 'include_debits' in request.POST
            
            # Parse the dates
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Validate date range
            if start_date > end_date:
                messages.error(request, 'Start date cannot be after end date')
                return redirect('report')
                
            # Get sales data for the date range
            range_sales = Sales.objects.filter(date_added__date__gte=start_date, date_added__date__lte=end_date)
            total_sales = range_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            total_sales = float(total_sales)
            
            # Get all week numbers that fall within this date range
            relevant_weeks = set()
            current_date = start_date
            while current_date <= end_date:
                year = current_date.year
                first_sunday_of_year = get_sunday(date(year, 1, 1))
                days_since_first_sunday = (current_date - first_sunday_of_year).days
                week_number = (days_since_first_sunday // 7) + 1
                relevant_weeks.add(week_number)
                current_date += timedelta(days=1)
            
            # Get all debits for these weeks
            debits = WeeklyDebit.objects.filter(week_number__in=relevant_weeks)
            total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
            total_debits = float(total_debits)
            
            # Calculate net sales
            net_sales = total_sales - total_debits
            
            # Calculate daily sales
            daily_sales = []
            current_date = start_date
            while current_date <= end_date:
                day_sales = Sales.objects.filter(date_added__date=current_date)
                day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
                day_total = float(day_total)
                
                daily_sales.append({
                    'date': current_date.strftime('%d-%b-%Y'),
                    'day': current_date.strftime('%a'),
                    'sales': day_total,
                })
                current_date += timedelta(days=1)
            
            # Create an Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Date Range Report"
            
            # Add title
            ws['A1'] = f"Sales Report - {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
            ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
            
            # Add date range
            date_range_days = (end_date - start_date).days + 1
            ws['A2'] = f"Period: {date_range_days} days"
            ws['A2'].font = openpyxl.styles.Font(italic=True)
            
            # Add summary
            ws['A4'] = "SUMMARY"
            ws['A4'].font = openpyxl.styles.Font(bold=True)
            
            ws['A5'] = "Gross Sales:"
            ws['B5'] = total_sales
            ws['B5'].number_format = '₹#,##0.00'
            
            ws['A6'] = "Total Debits:"
            ws['B6'] = total_debits
            ws['B6'].number_format = '₹#,##0.00'
            
            ws['A7'] = "Net Balance:"
            ws['B7'] = net_sales
            ws['B7'].font = openpyxl.styles.Font(bold=True)
            ws['B7'].number_format = '₹#,##0.00'
            
            row = 9  # Start row for the next section
            
            # Add daily sales if selected
            if include_daily_sales:
                ws[f'A{row}'] = "DAILY SALES"
                ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add headers
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Day"
                ws[f'C{row}'] = "Sales Amount"
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add data
                for day_data in daily_sales:
                    ws[f'A{row}'] = day_data['date']
                    ws[f'B{row}'] = day_data['day']
                    ws[f'C{row}'] = day_data['sales']
                    ws[f'C{row}'].number_format = '₹#,##0.00'
                    row += 1
                
                row += 2  # Add space after daily sales
            
            # Add debit history if selected
            if include_debits:
                ws[f'A{row}'] = "DEBIT HISTORY"
                ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                row += 1
                
                if debits.exists():
                    # Add headers
                    ws[f'A{row}'] = "Date"
                    ws[f'B{row}'] = "Week #"
                    ws[f'C{row}'] = "Description"
                    ws[f'D{row}'] = "Amount"
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.font = openpyxl.styles.Font(bold=True)
                    row += 1
                    
                    # Add data
                    for debit in debits:
                        ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                        ws[f'B{row}'] = f"Week {debit.week_number}"
                        ws[f'C{row}'] = debit.description
                        ws[f'D{row}'] = debit.amount
                        ws[f'D{row}'].number_format = '₹#,##0.00'
                        row += 1
                else:
                    ws[f'A{row}'] = "No debits recorded for this date range."
                    row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to a temporary file
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create the email
            email = EmailMessage(
                subject=f"Sales Report - {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}",
                body=f"Please find attached the sales report for the period {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}.\n\n"
                     f"Period: {date_range_days} days\n"
                     f"Gross Sales: ₹{total_sales:,.2f}\n"
                     f"Total Debits: ₹{total_debits:,.2f}\n"
                     f"Net Balance: ₹{net_sales:,.2f}",
                from_email="system@bavaastro.com",  # Update with your email
                to=[recipient_email]
            )
            
            # Attach the Excel file
            filename = f"Sales_Report_{start_date.strftime('%d_%b_%Y')}_to_{end_date.strftime('%d_%b_%Y')}.xlsx"
            email.attach(filename, excel_file.getvalue(), 
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            # Send the email
            try:
                email.send()
                messages.success(request, f'Sales report for {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")} has been emailed to {recipient_email}')
            except Exception as e:
                messages.error(request, f'Failed to send email: {str(e)}')
            
            return redirect('report')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid date format: {str(e)}')
            return redirect('report')
    
    messages.error(request, 'Invalid request method')
    return redirect('report')
    
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def email_weekly_report(request, week_number):
    if request.method == 'POST':
        include_daily_sales = 'include_daily_sales' in request.POST
        include_debits = 'include_debits' in request.POST
        recipient_email = request.POST.get('recipient_email', 'sivaprasanth8208@gmail.com')
        
        # Get data for the week
        today = timezone.now().date()
        first_sunday = get_sunday(date(today.year, 1, 1))
        week_start = first_sunday + timedelta(days=7 * (week_number - 1))
        week_end = week_start + timedelta(days=6)
        
        # Get sales data
        week_sales = Sales.objects.filter(date_added__date__gte=week_start, date_added__date__lte=week_end)
        total_sales = week_sales.aggregate(total=Sum('grand_total'))['total'] or 0
        total_sales = float(total_sales)
        
        # Get debit data
        debits = WeeklyDebit.objects.filter(week_number=week_number)
        total_debits = debits.aggregate(total=Sum('amount'))['total'] or 0
        total_debits = float(total_debits)
        
        # Calculate net sales
        net_sales = total_sales - total_debits
        
        # Calculate daily sales
        daily_sales = []
        current_date = week_start
        while current_date <= week_end:
            day_sales = Sales.objects.filter(date_added__date=current_date)
            day_total = day_sales.aggregate(total=Sum('grand_total'))['total'] or 0
            day_total = float(day_total)
            daily_sales.append({
                'date': current_date.strftime('%d-%b-%Y'),
                'day': current_date.strftime('%a'),
                'total': day_total
            })
            current_date += timedelta(days=1)
        
        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        
        # Add title
        ws['A1'] = f"Weekly Report - Week {week_number}"
        ws['A2'] = f"Period: {week_start.strftime('%d-%b-%Y')} to {week_end.strftime('%d-%b-%Y')}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A2'].font = openpyxl.styles.Font(italic=True)
        
        # Add summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = openpyxl.styles.Font(bold=True)
        
        ws['A5'] = "Total Gross Sales:"
        ws['B5'] = total_sales
        ws['B5'].number_format = '₹#,##0.00'
        
        ws['A6'] = "Total Debits:"
        ws['B6'] = total_debits
        ws['B6'].number_format = '₹#,##0.00'
        
        ws['A7'] = "Net Sales:"
        ws['B7'] = net_sales
        ws['B7'].font = openpyxl.styles.Font(bold=True)
        ws['B7'].number_format = '₹#,##0.00'
        
        row = 9  # Start row for the next section
        
        # Add daily sales if selected
        if include_daily_sales:
            ws[f'A{row}'] = "DAILY SALES"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add headers
            ws[f'A{row}'] = "Date"
            ws[f'B{row}'] = "Day"
            ws[f'C{row}'] = "Sales Amount"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add data
            for day in daily_sales:
                ws[f'A{row}'] = day['date']
                ws[f'B{row}'] = day['day']
                ws[f'C{row}'] = day['total']
                ws[f'C{row}'].number_format = '₹#,##0.00'
                
                # Highlight Tue, Wed, Thu as requested in the UI
                if day['day'] in ['Tue', 'Wed', 'Thu']:
                    for col in range(1, 4):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = openpyxl.styles.PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
                
                row += 1
            
            row += 2  # Add space after daily sales
        
        # Add debit history if selected
        if include_debits and debits.exists():
            ws[f'A{row}'] = "DEBIT HISTORY"
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add headers
            ws[f'A{row}'] = "Date"
            ws[f'B{row}'] = "Description"
            ws[f'C{row}'] = "Amount"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
            row += 1
            
            # Add data
            for debit in debits:
                ws[f'A{row}'] = debit.date_added.strftime('%d-%b-%Y')
                ws[f'B{row}'] = debit.description
                ws[f'C{row}'] = debit.amount
                ws[f'C{row}'].number_format = '₹#,##0.00'
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Prepare email
        subject = f"Weekly Report - Week {week_number} ({week_start.strftime('%d-%b-%Y')} to {week_end.strftime('%d-%b-%Y')})"
        body = f"""
        Dear User,
        
        Please find attached the weekly report for Week {week_number} ({week_start.strftime('%d-%b-%Y')} to {week_end.strftime('%d-%b-%Y')}).
        
        Summary:
        - Total Gross Sales: ₹{total_sales:,.2f}
        - Total Debits: ₹{total_debits:,.2f}
        - Net Sales: ₹{net_sales:,.2f}
        
        This is an automated email. Please do not reply.
        """
        
        filename = f"Weekly_Report_Week_{week_number}_{week_start.strftime('%d_%b_%Y')}.xlsx"
        
        # Send email with attachment
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=None,  # Uses DEFAULT_FROM_EMAIL from settings
            to=[recipient_email],
        )
        email.attach(filename, excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        try:
            email.send(fail_silently=False)
            messages.success(request, f"Report has been emailed to {recipient_email}")
        except Exception as e:
            messages.error(request, f"Failed to send email: {str(e)}")
        
        return redirect('weekly_report_detail', week_number=week_number)
    
    return redirect('weekly_report_detail', week_number=week_number)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def employee(request):
    user_groups = request.user.groups.all()
    u = request.user

    employees = Employee.objects.all()

    if request.method == "POST":
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        position = request.POST.get('position')
        daily_wage = request.POST.get('daily_wage')

        Employee.objects.create(
            name=name,
            phone_number=phone,
            position=position,
            daily_wage=daily_wage
        )
        return redirect('employee')   # reload page after saving

    context = {
        'page_title': 'Employee Details',
        'user_groups': user_groups,
        'u': u,
        'employees': employees,
    }
    return render(request, 'posApp/employee.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def edit_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == "POST":
        employee.name = request.POST.get('name')
        employee.phone_number = request.POST.get('phone')
        employee.position = request.POST.get('position')
        employee.daily_wage = request.POST.get('daily_wage')
        employee.save()
        return redirect('employee')

def delete_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    employee.delete()
    return redirect('employee')
    
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def mark_attendance(request, date):
    from posApp.models import Attendance
    
    attendance_date = datetime.strptime(date, '%Y-%m-%d').date()
    employees = Employee.objects.all()
    
    if request.method == 'POST':
        for employee in employees:
            present = request.POST.get(f'present_{employee.id}', '') == 'on'
            
            # Create or update attendance record
            Attendance.objects.update_or_create(
                employee=employee,
                date=attendance_date,
                defaults={'present': present}
            )
        
        messages.success(request, 'Attendance marked successfully')
        return redirect('attendance')
    
    # Get existing attendance records for this date
    attendance_records = {}
    for employee in employees:
        try:
            record = Attendance.objects.get(employee=employee, date=attendance_date)
            attendance_records[employee.id] = record.present
        except Attendance.DoesNotExist:
            attendance_records[employee.id] = False
    
    user_groups = request.user.groups.all()
    u = request.user
    
    context = {
        'page_title': f'Mark Attendance for {date}',
        'user_groups': user_groups,
        'u': u,
        'employees': employees,
        'attendance_date': attendance_date,
        'attendance_records': attendance_records,
    }
    return render(request, 'posApp/mark_attendance.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def attendance_summary(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from posApp.models import Attendance
    
    user_groups = request.user.groups.all()
    u = request.user
    
    # Get date range from request or default to current month
    today = timezone.now().date()
    
    # Default to first day of current month and today
    default_start = today.replace(day=1)
    default_end = today
    
    # Get start and end dates from request, or use defaults
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = default_start
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = default_end
    except ValueError:
        start_date = default_start
        end_date = default_end
    
    # Make sure start_date is not after end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date
        
    employees = Employee.objects.all()
    
    # Get all working days in the period (Tuesday, Wednesday, Thursday)
    all_working_days = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() in [1, 2, 3]:  # Tuesday, Wednesday, Thursday
            all_working_days.append(current_date)
        current_date += timedelta(days=1)
    
    # Calculate attendance statistics for each employee
    summary = {}
    total_wages = 0
    
    # Store detailed attendance records for each employee by date
    detailed_attendance = {}
    
    for employee in employees:
        # Get all attendance records for this employee in the date range
        records = Attendance.objects.filter(
            employee=employee,
            date__range=(start_date, end_date)
        )
        
        # Create a dictionary of date -> attendance record
        attendance_by_date = {record.date: record.present for record in records}
        detailed_attendance[employee] = attendance_by_date
        
        # Calculate working days and present days
        working_days = len(all_working_days)
        present_days = records.filter(present=True).count()
        
        # Calculate attendance percentage
        attendance_percentage = (present_days / working_days * 100) if working_days > 0 else 0
        
        # Calculate total wages
        total_wage = present_days * float(employee.daily_wage)
        total_wages += total_wage
        
        summary[employee] = {
            'working_days': working_days,
            'present_days': present_days,
            'attendance_percentage': attendance_percentage,
            'total_wages': total_wage
        }
    
    # Check if export action is requested
    if 'export' in request.GET:
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red
        not_marked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        
        # Add title and date range
        ws['A1'] = "Attendance and Wages Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Period: {start_date.strftime('%d %b, %Y')} to {end_date.strftime('%d %b, %Y')}"
        ws.merge_cells('A2:E2')
        
        # Add summary section header
        ws['A4'] = "ATTENDANCE SUMMARY"
        ws['A4'].font = Font(bold=True)
        ws.merge_cells('A4:E4')
        
        # Add summary headers
        headers = ["Employee Name", "Working Days", "Present Days", "Attendance %", "Total Wages (₹)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Add summary data
        row = 6
        for employee, data in summary.items():
            ws.cell(row=row, column=1, value=employee.name).border = thin_border
            ws.cell(row=row, column=2, value=data['working_days']).border = thin_border
            ws.cell(row=row, column=3, value=data['present_days']).border = thin_border
            ws.cell(row=row, column=4, value=f"{data['attendance_percentage']:.1f}%").border = thin_border
            ws.cell(row=row, column=5, value=float(data['total_wages'])).border = thin_border
            row += 1
        
        # Add totals row
        ws.cell(row=row, column=1, value="TOTAL").font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=5, value=float(total_wages)).font = header_font
        
        # Add space before detailed section
        row += 3
        
        # Add detailed section header
        ws.cell(row=row, column=1, value="DETAILED ATTENDANCE RECORD").font = Font(bold=True)
        ws.merge_cells(f'A{row}:' + get_column_letter(len(all_working_days) + 1) + f'{row}')
        row += 1
        
        # Add detailed headers
        ws.cell(row=row, column=1, value="Employee Name").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        
        # Add date headers
        for col, day in enumerate(all_working_days, start=2):
            cell = ws.cell(row=row, column=col, value=day.strftime("%a, %b %d"))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row += 1
        
        # Add detailed attendance data for each employee
        for employee in employees:
            ws.cell(row=row, column=1, value=employee.name).border = thin_border
            
            for col, day in enumerate(all_working_days, start=2):
                if day in detailed_attendance[employee]:
                    present = detailed_attendance[employee][day]
                    if present:
                        value = "Present"
                        fill = present_fill
                    else:
                        value = "Absent"
                        fill = absent_fill
                else:
                    value = "Not Marked"
                    fill = not_marked_fill
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.border = thin_border
            
            row += 1
        
        # Add daily wage details
        row += 2
        ws.cell(row=row, column=1, value="DAILY WAGES DETAILS").font = Font(bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Headers for daily wage table
        ws.cell(row=row, column=1, value="Employee").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value="Daily Wage").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value="Total Days Present").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value="Total Wages").font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        ws.cell(row=row, column=4).border = thin_border
        
        row += 1
        
        # Add daily wage data
        for employee in employees:
            ws.cell(row=row, column=1, value=employee.name).border = thin_border
            ws.cell(row=row, column=2, value=float(employee.daily_wage)).border = thin_border
            ws.cell(row=row, column=3, value=summary[employee]['present_days']).border = thin_border
            ws.cell(row=row, column=4, value=float(summary[employee]['total_wages'])).border = thin_border
            row += 1
            
        # Add grand total
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = header_font
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=4, value=float(total_wages)).font = header_font
        ws.cell(row=row, column=4).border = thin_border
        
        # Add legend
        row += 3
        ws.cell(row=row, column=1, value="Legend:").font = header_font
        row += 1
        
        # Present legend
        cell = ws.cell(row=row, column=1, value="Present")
        cell.fill = present_fill
        cell.border = thin_border
        row += 1
        
        # Absent legend
        cell = ws.cell(row=row, column=1, value="Absent")
        cell.fill = absent_fill
        cell.border = thin_border
        row += 1
        
        # Not marked legend
        cell = ws.cell(row=row, column=1, value="Not Marked")
        cell.fill = not_marked_fill
        cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(all_working_days) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Attendance_Report_{start_date}_to_{end_date}.xlsx'
        wb.save(response)
        return response
    
    # Check if email action is requested
    if 'email' in request.GET:
        recipient_email = request.GET.get('email_to', '')
        if not recipient_email:
            messages.error(request, 'Email address is required')
            return redirect(f'/attendance/summary/?start_date={start_date}&end_date={end_date}')
        
        # Create Excel file in memory
        import io
        output = io.BytesIO()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red
        not_marked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        
        # Add title and date range
        ws['A1'] = "Attendance and Wages Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Period: {start_date.strftime('%d %b, %Y')} to {end_date.strftime('%d %b, %Y')}"
        ws.merge_cells('A2:E2')
        
        # Add summary section header
        ws['A4'] = "ATTENDANCE SUMMARY"
        ws['A4'].font = Font(bold=True)
        ws.merge_cells('A4:E4')
        
        # Add headers
        headers = ["Employee Name", "Working Days", "Present Days", "Attendance %", "Total Wages (₹)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        row = 6
        for employee, data in summary.items():
            ws.cell(row=row, column=1, value=employee.name)
            ws.cell(row=row, column=2, value=data['working_days'])
            ws.cell(row=row, column=3, value=data['present_days'])
            ws.cell(row=row, column=4, value=f"{data['attendance_percentage']:.1f}%")
            ws.cell(row=row, column=5, value=float(data['total_wages']))
            row += 1
        
        # Add totals row
        ws.cell(row=row, column=1, value="TOTAL").font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=5, value=float(total_wages)).font = header_font
        
        # Add space before detailed section
        row += 3
        
        # Add detailed section header
        ws.cell(row=row, column=1, value="DETAILED ATTENDANCE RECORD").font = Font(bold=True)
        ws.merge_cells(f'A{row}:' + get_column_letter(len(all_working_days) + 1) + f'{row}')
        row += 1
        
        # Add detailed headers
        ws.cell(row=row, column=1, value="Employee Name").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        # Add date headers
        for col, day in enumerate(all_working_days, start=2):
            cell = ws.cell(row=row, column=col, value=day.strftime("%a, %b %d"))
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        
        # Add detailed attendance data for each employee
        for employee in employees:
            ws.cell(row=row, column=1, value=employee.name)
            
            for col, day in enumerate(all_working_days, start=2):
                if day in detailed_attendance[employee]:
                    present = detailed_attendance[employee][day]
                    if present:
                        value = "Present"
                        fill = present_fill
                    else:
                        value = "Absent"
                        fill = absent_fill
                else:
                    value = "Not Marked"
                    fill = not_marked_fill
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
            
            row += 1
        
        # Add daily wage details
        row += 2
        ws.cell(row=row, column=1, value="DAILY WAGES DETAILS").font = Font(bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Headers for daily wage table
        ws.cell(row=row, column=1, value="Employee").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2, value="Daily Wage").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3, value="Total Days Present").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        
        ws.cell(row=row, column=4, value="Total Wages").font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        
        row += 1
        
        # Add daily wage data
        for employee in employees:
            ws.cell(row=row, column=1, value=employee.name)
            ws.cell(row=row, column=2, value=float(employee.daily_wage))
            ws.cell(row=row, column=3, value=summary[employee]['present_days'])
            ws.cell(row=row, column=4, value=float(summary[employee]['total_wages']))
            row += 1
            
        # Add grand total
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = header_font
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=4, value=float(total_wages)).font = header_font
        
        # Add legend
        row += 3
        ws.cell(row=row, column=1, value="Legend:").font = header_font
        row += 1
        
        # Present legend
        cell = ws.cell(row=row, column=1, value="Present")
        cell.fill = present_fill
        row += 1
        
        # Absent legend
        cell = ws.cell(row=row, column=1, value="Absent")
        cell.fill = absent_fill
        row += 1
        
        # Not marked legend
        cell = ws.cell(row=row, column=1, value="Not Marked")
        cell.fill = not_marked_fill
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(all_working_days) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to BytesIO
        wb.save(output)
        
        # Send email
        try:
            email = EmailMessage(
                subject=f'Attendance Report {start_date} to {end_date}',
                body=f'Please find attached the detailed attendance report for the period {start_date} to {end_date}.',
                from_email=None,  # Uses DEFAULT_FROM_EMAIL from settings
                to=[recipient_email]
            )
            
            output.seek(0)  # Move to the beginning of the BytesIO object
            email.attach(f'Attendance_Report_{start_date}_to_{end_date}.xlsx', output.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()
            
            messages.success(request, f'Report successfully sent to {recipient_email}')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')
        
        return redirect(f'/attendance/summary/?start_date={start_date}&end_date={end_date}')
    
    context = {
        'page_title': 'Attendance Summary',
        'user_groups': user_groups,
        'u': u,
        'start_date': start_date,
        'end_date': end_date,
        'summary': summary,
        'total_wages': total_wages
    }
    
    return render(request, 'posApp/attendance_summary.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def attendance(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    
    user_groups = request.user.groups.all()
    u = request.user
    
    employees = Employee.objects.all()
    
    # Get selected week from request or default to current week
    selected_week = request.GET.get('week', '0')
    weeks_offset = int(selected_week)
    
    # Get current week dates (Tuesday, Wednesday, Thursday)
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    
    # Adjust week start based on selected week
    selected_week_start = current_week_start - timedelta(weeks=weeks_offset)
    
    # Filter days to just Tuesday, Wednesday, and Thursday
    working_days = []
    for i in range(7):
        day = selected_week_start + timedelta(days=i)
        # 1 = Tuesday, 2 = Wednesday, 3 = Thursday
        if day.weekday() in [1, 2, 3]:
            working_days.append(day)
    
    # Create a list of available weeks for selection (current and 4 previous weeks)
    available_weeks = []
    for i in range(5):  # Current week and 4 previous weeks
        week_start = current_week_start - timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        week_label = f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d')}"
        if i == 0:
            week_label += " (Current)"
        available_weeks.append({
            'offset': i,
            'label': week_label,
            'start': week_start,
            'end': week_end
        })
    
    # Handle attendance marking
    if request.method == 'POST':
        from posApp.models import Attendance
        
        date_str = request.POST.get('attendance_date')
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        for employee in employees:
            present = request.POST.get(f'present_{employee.id}', '') == 'on'
            
            # Create or update attendance record
            Attendance.objects.update_or_create(
                employee=employee,
                date=attendance_date,
                defaults={'present': present}
            )
        
        messages.success(request, 'Attendance marked successfully')
        return redirect('attendance')
    
    # Get attendance records for the selected week
    from posApp.models import Attendance
    attendance_records = {}
    
    for employee in employees:
        attendance_records[employee.id] = {}
        for day in working_days:
            try:
                record = Attendance.objects.get(employee=employee, date=day)
                attendance_records[employee.id][day] = record
            except Attendance.DoesNotExist:
                attendance_records[employee.id][day] = None
    
    # Calculate weekly wages for each employee
    weekly_wages = {}
    for employee in employees:
        total_wage = 0
        for day in working_days:
            if day in attendance_records[employee.id] and attendance_records[employee.id][day] and attendance_records[employee.id][day].present:
                total_wage += float(employee.daily_wage)
        weekly_wages[employee.id] = total_wage
    
    # Handle export request
    if 'export' in request.GET:
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Attendance"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Add title
        week_start = selected_week_start
        week_end = week_start + timedelta(days=6)
        ws['A1'] = f"Weekly Attendance: {week_start.strftime('%b %d, %Y')} to {week_end.strftime('%b %d, %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:' + get_column_letter(len(working_days) + 2) + '1')
        
        # Add headers
        ws.cell(row=3, column=1, value="Employee Name").font = header_font
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).border = thin_border
        
        # Day headers
        for col, day in enumerate(working_days, start=2):
            cell = ws.cell(row=3, column=col, value=day.strftime("%a, %b %d"))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Wages header
        ws.cell(row=3, column=len(working_days) + 2, value="Weekly Wage").font = header_font
        ws.cell(row=3, column=len(working_days) + 2).fill = header_fill
        ws.cell(row=3, column=len(working_days) + 2).border = thin_border
        
        # Add employee data
        row = 4
        for employee in employees:
            # Employee name
            ws.cell(row=row, column=1, value=employee.name).border = thin_border
            
            # Attendance for each day
            for col, day in enumerate(working_days, start=2):
                record = attendance_records[employee.id].get(day)
                if record:
                    if record.present:
                        value = "Present"
                        fill = present_fill
                    else:
                        value = "Absent"
                        fill = absent_fill
                else:
                    value = "Not Marked"
                    fill = None
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
            
            # Weekly wage
            ws.cell(row=row, column=len(working_days) + 2, value=weekly_wages[employee.id]).border = thin_border
            
            row += 1
        
        # Add total row
        ws.cell(row=row, column=1, value="TOTAL").font = header_font
        ws.merge_cells(f'A{row}:{get_column_letter(len(working_days) + 1)}{row}')
        total_cell = ws.cell(row=row, column=len(working_days) + 2, value=sum(weekly_wages.values()))
        total_cell.font = header_font
        total_cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        for i in range(len(working_days)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 15
        ws.column_dimensions[get_column_letter(len(working_days) + 2)].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Weekly_Attendance_{week_start.strftime("%Y%m%d")}_to_{week_end.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
    
    # Handle email request
    if 'email' in request.GET:
        import io
        recipient_email = request.GET.get('email_to', '')
        if not recipient_email:
            messages.error(request, 'Email address is required')
            return redirect(f'/attendance/?week={weeks_offset}')
        
        # Create Excel file in memory
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Attendance"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red
        not_marked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Add title
        week_start = selected_week_start
        week_end = week_start + timedelta(days=6)
        ws['A1'] = f"Weekly Attendance: {week_start.strftime('%b %d, %Y')} to {week_end.strftime('%b %d, %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:' + get_column_letter(len(working_days) + 2) + '1')
        
        # Add summary section header
        ws['A3'] = "ATTENDANCE SUMMARY"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells(f'A3:{get_column_letter(len(working_days) + 2)}3')
        
        # Add headers
        ws.cell(row=5, column=1, value="Employee Name").font = header_font
        ws.cell(row=5, column=1).fill = header_fill
        ws.cell(row=5, column=1).border = thin_border
        
        # Day headers
        for col, day in enumerate(working_days, start=2):
            cell = ws.cell(row=5, column=col, value=day.strftime("%a, %b %d"))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            # Set the column width
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Wages header
        wages_col = len(working_days) + 2
        ws.cell(row=5, column=wages_col, value="Weekly Wage").font = header_font
        ws.cell(row=5, column=wages_col).fill = header_fill
        ws.cell(row=5, column=wages_col).border = thin_border
        ws.column_dimensions[get_column_letter(wages_col)].width = 15
        
        # Add employee data
        row = 6
        for employee in employees:
            # Employee name
            name_cell = ws.cell(row=row, column=1, value=employee.name)
            name_cell.border = thin_border
            
            # Attendance for each day
            for col, day in enumerate(working_days, start=2):
                record = attendance_records[employee.id].get(day)
                if record:
                    if record.present:
                        value = "Present"
                        fill = present_fill
                    else:
                        value = "Absent"
                        fill = absent_fill
                else:
                    value = "Not Marked"
                    fill = not_marked_fill
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.border = thin_border
            
            # Weekly wage
            wage_cell = ws.cell(row=row, column=wages_col, value=weekly_wages[employee.id])
            wage_cell.border = thin_border
            
            row += 1
        
        # Add total row
        total_label = ws.cell(row=row, column=1, value="TOTAL")
        total_label.font = header_font
        total_label.border = thin_border
        
        ws.merge_cells(f'A{row}:{get_column_letter(len(working_days) + 1)}{row}')
        
        total_cell = ws.cell(row=row, column=wages_col, value=sum(weekly_wages.values()))
        total_cell.font = header_font
        total_cell.border = thin_border
        
        # Add detailed wage section
        row += 3
        ws.cell(row=row, column=1, value="DAILY WAGES DETAILS").font = Font(bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Headers for daily wage table
        ws.cell(row=row, column=1, value="Employee").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value="Daily Wage").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value="Days Present").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value="Total Wage").font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        ws.cell(row=row, column=4).border = thin_border
        
        row += 1
        
        # Add employee wage details
        for employee in employees:
            # Count present days
            present_days = sum(1 for day in working_days if 
                              day in attendance_records[employee.id] and 
                              attendance_records[employee.id][day] and 
                              attendance_records[employee.id][day].present)
            
            ws.cell(row=row, column=1, value=employee.name).border = thin_border
            ws.cell(row=row, column=2, value=float(employee.daily_wage)).border = thin_border
            ws.cell(row=row, column=3, value=present_days).border = thin_border
            ws.cell(row=row, column=4, value=float(weekly_wages[employee.id])).border = thin_border
            row += 1
            
        # Add total row
        ws.cell(row=row, column=1, value="TOTAL").font = header_font
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(f'A{row}:C{row}')
        
        for col in range(2, 4):
            ws.cell(row=row, column=col).border = thin_border
            
        ws.cell(row=row, column=4, value=float(sum(weekly_wages.values()))).font = header_font
        ws.cell(row=row, column=4).border = thin_border
        
        # Add legend
        row += 3
        ws.cell(row=row, column=1, value="Legend:").font = header_font
        row += 1
        
        # Present legend
        cell = ws.cell(row=row, column=1, value="Present")
        cell.fill = present_fill
        cell.border = thin_border
        row += 1
        
        # Absent legend
        cell = ws.cell(row=row, column=1, value="Absent")
        cell.fill = absent_fill
        cell.border = thin_border
        row += 1
        
        # Not marked legend
        cell = ws.cell(row=row, column=1, value="Not Marked")
        cell.fill = not_marked_fill
        cell.border = thin_border
        
        # Set column widths for names
        ws.column_dimensions['A'].width = 20
        
        # Save to BytesIO
        wb.save(output)
        
        # Send email
        try:
            email = EmailMessage(
                subject=f'Weekly Attendance Report: {week_start.strftime("%b %d")} to {week_end.strftime("%b %d")}',
                body=f'Please find attached the detailed weekly attendance report for the period {week_start.strftime("%b %d, %Y")} to {week_end.strftime("%b %d, %Y")}.',
                from_email=None,  # Uses DEFAULT_FROM_EMAIL from settings
                to=[recipient_email]
            )
            
            output.seek(0)  # Move to the beginning of the BytesIO object
            filename = f'Weekly_Attendance_{week_start.strftime("%Y%m%d")}_to_{week_end.strftime("%Y%m%d")}.xlsx'
            email.attach(filename, output.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()
            
            messages.success(request, f'Report successfully sent to {recipient_email}')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')
        
        return redirect(f'/attendance/?week={weeks_offset}')
    
    context = {
        'page_title': 'Attendance Details',
        'user_groups': user_groups,
        'u': u,
        'employees': employees,
        'working_days': working_days,
        'attendance_records': attendance_records,
        'weekly_wages': weekly_wages,
        'available_weeks': available_weeks,
        'selected_week': weeks_offset,
    }
    return render(request, 'posApp/attendance.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def bulk_upload(request):
    user_groups = request.user.groups.all()
    u = request.user
    
    context = {
        'page_title': 'Bulk Upload/Export',
        'user_groups': user_groups,
        'u': u,
    }
    return render(request, 'posApp/bulk_upload.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def export_bulk_data(request):
    """Export all data from models as JSON"""
    try:
        data = {
            'categories': [],
            'products': [],
            'employees': [],
            'customers': [],
            'sales': [],
            'sales_items': [],
            'weekly_debits': [],
            'attendance': [],
        }
        
        # Export Categories
        for category in Category.objects.all():
            data['categories'].append({
                'id': category.id,
                'name': category.name,
                'status': category.status,
                'date_added': category.date_added.isoformat(),
                'date_updated': category.date_updated.isoformat(),
            })
        
        # Export Products
        for product in Products.objects.all():
            data['products'].append({
                'id': product.id,
                'category_id': product.category_id.id,
                'name': product.name,
                'price': product.price,
                'status': product.status,
                'date_added': product.date_added.isoformat(),
                'date_updated': product.date_updated.isoformat(),
            })
        
        # Export Employees
        for employee in Employee.objects.all():
            data['employees'].append({
                'id': employee.id,
                'name': employee.name,
                'phone_number': employee.phone_number,
                'position': employee.position,
                'daily_wage': float(employee.daily_wage),
            })
        
        # Export Customers
        for customer in Customer.objects.all():
            data['customers'].append({
                'id': customer.id,
                'name': customer.name,
                'phone_number': customer.phone_number,
                'city': customer.city,
            })
        
        # Export Sales
        for sale in Sales.objects.all():
            data['sales'].append({
                'id': sale.id,
                'customer_name': sale.customer_name,
                'customer_phone': sale.customer_phone,
                'customer_city': sale.customer_city,
                'payment_method': sale.payment_method,
                'sub_total': sale.sub_total,
                'grand_total': sale.grand_total,
                'room_no': sale.room_no,
                'date_added': sale.date_added.isoformat(),
                'token_no': sale.token_no,
                'raw_token_no': sale.raw_token_no,
                'serial_no': sale.serial_no,
            })
        
        # Export Sales Items
        for item in salesItems.objects.all():
            data['sales_items'].append({
                'id': item.id,
                'sale_id': item.sale_id.id,
                'product_id': item.product_id.id,
                'price': item.price,
                'qty': item.qty,
                'total': item.total,
            })
        
        # Export Weekly Debits
        for debit in WeeklyDebit.objects.all():
            data['weekly_debits'].append({
                'id': debit.id,
                'week_number': debit.week_number,
                'week_start_date': debit.week_start_date.isoformat(),
                'description': debit.description,
                'amount': float(debit.amount),
                'date_added': debit.date_added.isoformat(),
            })
        
        # Export Attendance
        for attendance in Attendance.objects.all():
            data['attendance'].append({
                'id': attendance.id,
                'employee_id': attendance.employee.id,
                'date': attendance.date.isoformat(),
                'present': attendance.present,
                'date_added': attendance.date_added.isoformat(),
            })
        
        # Create response
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        export_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="pos_data_export_{export_date}.json"'
        return response
        
    except Exception as e:
        return JsonResponse({
            'status': 'failed',
            'msg': f'Export failed: {str(e)}'
        }, status=400)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@csrf_exempt
def import_bulk_data(request):
    """Import data from JSON file and save to database"""
    try:
        # Check authentication
        if not request.user.is_authenticated:
            return JsonResponse({'status': 'failed', 'msg': 'Authentication required'}, status=401)
        
        if request.method != 'POST':
            return JsonResponse({'status': 'failed', 'msg': 'Invalid request method'}, status=400)
        
        if 'file' not in request.FILES:
            return JsonResponse({'status': 'failed', 'msg': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        
        if not file.name.endswith('.json'):
            return JsonResponse({'status': 'failed', 'msg': 'Only JSON files are allowed'}, status=400)
        
        # Read and parse JSON
        file_content = file.read().decode('utf-8')
        data = json.loads(file_content)
        
        # Track imported records
        imported_count = {
            'categories': 0,
            'products': 0,
            'employees': 0,
            'customers': 0,
            'sales': 0,
            'sales_items': 0,
            'weekly_debits': 0,
            'attendance': 0,
        }
        
        errors = []
        
        # Import Categories
        try:
            for item in data.get('categories', []):
                try:
                    Category.objects.update_or_create(
                        id=item['id'],
                        defaults={
                            'name': item['name'],
                            'status': item['status'],
                            'date_added': item['date_added'],
                            'date_updated': item['date_updated'],
                        }
                    )
                    imported_count['categories'] += 1
                except Exception as item_error:
                    errors.append(f'Category ID {item.get("id", "unknown")} error: {str(item_error)}')
        except Exception as e:
            errors.append(f'Category import error: {str(e)}')
        
        # Import Products
        try:
            for item in data.get('products', []):
                try:
                    category = Category.objects.get(id=item['category_id'])
                    Products.objects.update_or_create(
                        id=item['id'],
                        defaults={
                            'category_id': category,
                            'name': item['name'],
                            'price': item['price'],
                            'status': item['status'],
                            'date_added': item['date_added'],
                            'date_updated': item['date_updated'],
                        }
                    )
                    imported_count['products'] += 1
                except Category.DoesNotExist:
                    errors.append(f'Product ID {item.get("id", "unknown")}: Category {item.get("category_id")} not found')
                except Exception as item_error:
                    errors.append(f'Product ID {item.get("id", "unknown")} error: {str(item_error)}')
        except Exception as e:
            errors.append(f'Product import error: {str(e)}')
        
        # Import Employees
        try:
            for item in data.get('employees', []):
                Employee.objects.update_or_create(
                    id=item['id'],
                    defaults={
                        'name': item['name'],
                        'phone_number': item['phone_number'],
                        'position': item['position'],
                        'daily_wage': item['daily_wage'],
                    }
                )
                imported_count['employees'] += 1
        except Exception as e:
            errors.append(f'Employee import error: {str(e)}')
        
        # Import Customers
        try:
            for item in data.get('customers', []):
                Customer.objects.update_or_create(
                    id=item['id'],
                    defaults={
                        'name': item['name'],
                        'phone_number': item['phone_number'],
                        'city': item.get('city', ''),
                    }
                )
                imported_count['customers'] += 1
        except Exception as e:
            errors.append(f'Customer import error: {str(e)}')
        
        # Import Sales
        try:
            for item in data.get('sales', []):
                try:
                    Sales.objects.update_or_create(
                        id=item['id'],
                        defaults={
                            'customer_name': item['customer_name'],
                            'customer_phone': item['customer_phone'],
                            'customer_city': item['customer_city'],
                            'payment_method': item['payment_method'],
                            'sub_total': item['sub_total'],
                            'grand_total': item['grand_total'],
                            'room_no': item['room_no'],
                            'token_no': item['token_no'],
                            'raw_token_no': item.get('raw_token_no'),
                            'serial_no': item.get('serial_no', 1),
                            'date_added': item['date_added'],
                        }
                    )
                    imported_count['sales'] += 1
                except Exception as item_error:
                    errors.append(f'Sale ID {item.get("id", "unknown")} error: {str(item_error)}')
        except Exception as e:
            errors.append(f'Sales import error: {str(e)}')
        
        # Import Sales Items
        try:
            for item in data.get('sales_items', []):
                try:
                    sale = Sales.objects.get(id=item['sale_id'])
                    product = Products.objects.get(id=item['product_id'])
                    salesItems.objects.update_or_create(
                        id=item['id'],
                        defaults={
                            'sale_id': sale,
                            'product_id': product,
                            'price': item['price'],
                            'qty': item['qty'],
                            'total': item['total'],
                        }
                    )
                    imported_count['sales_items'] += 1
                except Sales.DoesNotExist:
                    errors.append(f'Sales Item ID {item.get("id", "unknown")}: Sale {item.get("sale_id")} not found')
                except Products.DoesNotExist:
                    errors.append(f'Sales Item ID {item.get("id", "unknown")}: Product {item.get("product_id")} not found')
                except Exception as item_error:
                    errors.append(f'Sales Item ID {item.get("id", "unknown")} error: {str(item_error)}')
        except Exception as e:
            errors.append(f'Sales Items import error: {str(e)}')
        
        # Import Weekly Debits
        try:
            for item in data.get('weekly_debits', []):
                WeeklyDebit.objects.update_or_create(
                    id=item['id'],
                    defaults={
                        'week_number': item['week_number'],
                        'week_start_date': item['week_start_date'],
                        'description': item['description'],
                        'amount': item['amount'],
                        'date_added': item['date_added'],
                    }
                )
                imported_count['weekly_debits'] += 1
        except Exception as e:
            errors.append(f'Weekly Debit import error: {str(e)}')
        
        # Import Attendance
        try:
            for item in data.get('attendance', []):
                try:
                    employee = Employee.objects.get(id=item['employee_id'])
                    Attendance.objects.update_or_create(
                        id=item['id'],
                        defaults={
                            'employee': employee,
                            'date': item['date'],
                            'present': item['present'],
                            'date_added': item['date_added'],
                        }
                    )
                    imported_count['attendance'] += 1
                except Employee.DoesNotExist:
                    errors.append(f'Attendance ID {item.get("id", "unknown")}: Employee {item.get("employee_id")} not found')
                except Exception as item_error:
                    errors.append(f'Attendance ID {item.get("id", "unknown")} error: {str(item_error)}')
        except Exception as e:
            errors.append(f'Attendance import error: {str(e)}')
        
        # Recalculate serial numbers for all sales to maintain correct order
        try:
            all_sales = Sales.objects.all().order_by('id')

            for index, sale in enumerate(all_sales, start=1):
                sale.serial_no = index
                sale.save(update_fields=['serial_no'])
        except Exception as e:
            errors.append(f'Serial number recalculation error: {str(e)}')
        
        return JsonResponse({
            'status': 'success',
            'msg': 'Data imported successfully',
            'imported': imported_count,
            'errors': errors if errors else None,
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'failed', 'msg': 'Invalid JSON file'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'failed', 'msg': f'Import failed: {str(e)}'}, status=400)

        